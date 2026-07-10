# Copyright (c) 2016. Mount Sinai School of Medicine
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""ERAMER — ERAP1 trimming prediction, filling mhctools' `erap_trimming` gap.

ERAP1 (ER aminopeptidase 1) trims the N-termini of 9-16mer precursor peptides in
the ER down to the 8-10mers that MHC-I presents — the processing step between
TAP transport and MHC loading, and previously the one empty ``Kind`` in
mhctools. ERAMER (Al-okaily 2024) models ERAP1 specificity as a per-length
position-weight matrix and scores a precursor by averaging the PWM specificity
over each residue trimmed off as it is cut down to a target epitope length.

Licensing: ERAMER is GPLv3 and its PWM ships in a GPL-licensed ``PWM.xlsx``;
mhctools is Apache-2.0 and vendors neither. This is a clean-room Python-3
reimplementation of the trimming-cascade average (the upstream tool is Python
2.7) that loads the PWM from a user-provided ERAMER checkout at runtime, as the
netMHC / MixMHCpred wrappers read user-provided model files. Point at the
checkout with ``ERAMER_HOME`` (or pass ``eramer_home=`` / ``pwm_path=``).

Upstream: https://github.com/aalokaily/ERAMER
Cite: Al-okaily et al., Comput. Biol. Med. 2024 — "ERAMER: A novel in silico
tool for prediction of ERAP1 enzyme trimming".

ERAMER's evaluation is self-reported and ERAP1 trimming is intrinsically noisy —
treat the score (roughly -1..1, higher = more likely trimmed) as a pathway
prior, not a validated oracle.
"""

import os
from os.path import isdir, isfile, join

from .pred import Kind, PeptideResult, Prediction
from .wrapper_base import AlleleFreePredictor

# ERAP1 processes precursors in this length range; ERAMER ships one PWM sheet
# per precursor length in [9, 16].
ERAMER_MIN_PRECURSOR_LENGTH = 9
ERAMER_MAX_PRECURSOR_LENGTH = 16
_PWM_LENGTHS = tuple(range(
    ERAMER_MIN_PRECURSOR_LENGTH, ERAMER_MAX_PRECURSOR_LENGTH + 1))
_VALID_AMINO_ACIDS = frozenset("ACDEFGHIKLMNPQRSTVWY")


_MISSING_PWM_HELP = (
    "Set ERAMER_HOME to a clone of https://github.com/aalokaily/ERAMER "
    "(or pass eramer_home= / pwm_path=). mhctools does not vendor the "
    "GPL-licensed PWM.")


def _find_pwm_path(eramer_home=None, pwm_path=None):
    """Resolve the path to ERAMER's ``PWM.xlsx``.

    An explicit *pwm_path* or *eramer_home* is honored exactly (and must exist —
    no silent fallback). Otherwise the resolver tries ``$ERAMER_PWM``, then
    ``$ERAMER_HOME/PWM.xlsx``, then ``~/ERAMER/PWM.xlsx``.
    """
    if pwm_path:
        if isfile(pwm_path):
            return pwm_path
        raise FileNotFoundError(
            "ERAMER PWM not found at pwm_path=%r." % pwm_path)
    if eramer_home:
        candidate = join(eramer_home, "PWM.xlsx")
        if isfile(candidate):
            return candidate
        raise FileNotFoundError(
            "PWM.xlsx not found in eramer_home=%r." % eramer_home)

    candidates = []
    if os.environ.get("ERAMER_PWM"):
        candidates.append(os.environ["ERAMER_PWM"])
    if os.environ.get("ERAMER_HOME"):
        candidates.append(join(os.environ["ERAMER_HOME"], "PWM.xlsx"))
    home = join(os.path.expanduser("~"), "ERAMER")
    if isdir(home):
        candidates.append(join(home, "PWM.xlsx"))
    for candidate in candidates:
        if isfile(candidate):
            return candidate
    raise FileNotFoundError("ERAMER's PWM.xlsx not found. " + _MISSING_PWM_HELP)


def load_pwm(pwm_path):
    """Load ERAMER's PWM into ``{length: {amino_acid: [weight_per_position]}}``.

    Reads the ``Length 9`` .. ``Length 16`` sheets: each has 20 amino-acid rows
    (letter in column B, per-position weights in the following columns).
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError(
            "Reading ERAMER's PWM.xlsx requires openpyxl (`pip install "
            "openpyxl`).") from e

    workbook = openpyxl.load_workbook(pwm_path, data_only=True, read_only=True)
    weights_by_length = {}
    for length in _PWM_LENGTHS:
        sheet_name = "Length %d" % length
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                "ERAMER PWM.xlsx is missing sheet %r" % sheet_name)
        worksheet = workbook[sheet_name]
        aa_weights = {}
        # 20 amino-acid rows: openpyxl rows 2..21, letter in col 2, the L
        # per-position weights in cols 3..(2 + length).
        for row in range(2, 22):
            aa = worksheet.cell(row=row, column=2).value
            if not aa:
                continue
            aa = str(aa).strip().upper()
            aa_weights[aa] = [
                float(worksheet.cell(row=row, column=3 + pos).value)
                for pos in range(length)]
        missing = _VALID_AMINO_ACIDS - set(aa_weights)
        if missing:
            raise ValueError(
                "ERAMER PWM sheet %r missing amino acid(s): %s"
                % (sheet_name, "".join(sorted(missing))))
        weights_by_length[length] = aa_weights
    workbook.close()
    return weights_by_length


def _specificity(peptide, aa_weights):
    """Mean PWM specificity of a peptide (one trimming intermediate)."""
    length = len(peptide)
    total = sum(aa_weights[peptide[i]][i] for i in range(length))
    # Rounded per-residue-average, matching upstream ERAMER's compute_specificty.
    return round(total / length, 6)


def eramer_score(precursor, epitope_length, weights_by_length):
    """ERAMER trimming score: mean specificity over the trimming cascade.

    As the precursor is trimmed one residue at a time from length ``L`` down to
    ``epitope_length + 1``, each intermediate C-terminal ``l``-mer is scored by
    its PWM specificity; the returned score is their mean (higher = more likely
    trimmed by ERAP1).

    Returns ``None`` if the precursor is already at/below ``epitope_length``
    (nothing to trim).
    """
    length = len(precursor)
    scores = [
        _specificity(precursor[-sub_len:], weights_by_length[sub_len])
        for sub_len in range(length, epitope_length, -1)]
    if not scores:
        return None
    return sum(scores) / len(scores)


class ERAMER(AlleleFreePredictor):
    """ERAMER ERAP1-trimming predictor (clean-room reimplementation).

    Allele-independent: ``predict()`` returns one ``Kind.erap_trimming``
    prediction per peptide (the peptide is treated as an ERAP1 *precursor*),
    with an empty ``allele``.

    Parameters
    ----------
    epitope_length : int
        Target epitope length the precursor is trimmed toward; the cascade runs
        from the precursor length down to ``epitope_length + 1``. Default 8 (so
        every valid 9-16mer precursor yields a score). Must be in 8-15.
    eramer_home : str, optional
        Path to an ERAMER checkout (containing ``PWM.xlsx``). If omitted, the PWM
        is resolved from ``$ERAMER_PWM``, then ``$ERAMER_HOME``, then
        ``~/ERAMER``.
    pwm_path : str, optional
        Direct path to ``PWM.xlsx`` (overrides *eramer_home*).
    """

    mhc_class = "I"

    def __init__(self, epitope_length=8, eramer_home=None, pwm_path=None):
        if not (8 <= epitope_length <= ERAMER_MAX_PRECURSOR_LENGTH - 1):
            raise ValueError(
                "epitope_length must be in 8-%d (the shortest trimmed "
                "intermediate must still be a >=9mer with a PWM); got %d"
                % (ERAMER_MAX_PRECURSOR_LENGTH - 1, epitope_length))
        self.epitope_length = epitope_length
        self.pwm_path = _find_pwm_path(eramer_home, pwm_path)
        self._weights_by_length = None

    def __str__(self):
        return "ERAMER(epitope_length=%d, pwm_path=%r)" % (
            self.epitope_length, self.pwm_path)

    def _default_pred_kind(self):
        return Kind.erap_trimming

    def _ensure_loaded(self):
        if self._weights_by_length is None:
            self._weights_by_length = load_pwm(self.pwm_path)
        return self._weights_by_length

    def _check_peptides(self, peptides):
        for peptide in peptides:
            n = len(peptide)
            if not (ERAMER_MIN_PRECURSOR_LENGTH <= n
                    <= ERAMER_MAX_PRECURSOR_LENGTH):
                raise ValueError(
                    "ERAMER precursors must be %d-%d residues; got %r (length "
                    "%d)" % (ERAMER_MIN_PRECURSOR_LENGTH,
                             ERAMER_MAX_PRECURSOR_LENGTH, peptide, n))
            if n <= self.epitope_length:
                raise ValueError(
                    "Precursor %r (length %d) is not longer than the target "
                    "epitope_length %d, so there is nothing to trim"
                    % (peptide, n, self.epitope_length))
            invalid = set(peptide) - _VALID_AMINO_ACIDS
            if invalid:
                raise ValueError(
                    "Peptide %r contains non-standard amino acid(s): %s"
                    % (peptide, "".join(sorted(invalid))))

    def predict(self, peptides):
        """Predict ERAP1 trimming for a list of precursor peptides.

        Returns
        -------
        list of PeptideResult
            One entry per input peptide, each holding a single
            ``Kind.erap_trimming`` prediction (empty ``allele``; ``score`` is the
            ERAMER trimming score, higher = more likely trimmed).
        """
        peptide_list = self._normalize_peptides(peptides)
        self._check_peptides(peptide_list)
        if not peptide_list:
            return []
        weights_by_length = self._ensure_loaded()

        # _check_peptides guarantees each precursor is longer than
        # epitope_length, so eramer_score always yields a value here.
        return [
            PeptideResult(preds=(Prediction(
                kind=Kind.erap_trimming,
                score=float(eramer_score(
                    peptide, self.epitope_length, weights_by_length)),
                peptide=peptide,
                predictor_name="eramer"),))
            for peptide in peptide_list
        ]
